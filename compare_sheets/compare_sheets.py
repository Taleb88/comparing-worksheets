import pandas as pd

# RAW FILE

df = pd.read_excel('raw_file.xlsx')
df['Media'].replace('VOD', 'TVOD', inplace=True) # replace media column values = 'vod' with 'tvod'

# sort column values by 'name', 'media', and 'resolution'
x = df.sort_values(by=['Name', 'Media', 'Resolution'], axis=0, ascending=True)
# print(x)

# select only 'name' and 'media' columns to be copied to new workbook -> result.xlsx
new_df = pd.DataFrame()
new_df['Name'] = x['Name'].copy()
new_df['Media'] = x['Media'].copy()
new_df['Resolution'] = x['Resolution'].copy()

# print(new_df)
new_df.to_excel('raw_file_condensed.xlsx') # create condensed version of raw file


# EXTRACT

extract = pd.read_excel('extract.xlsx')

# sort column values by 'name', 'media', and 'resolution'
y = extract.sort_values(by=['Name', 'Media', 'Resolution'], axis=0, ascending=True)

# print(y)

# pull 'name', 'media', and 'resolution' columns from extract
#   by selecting the aforementioned columns to be copied to new workbook
new_df2 = pd.DataFrame()
new_df2['Name'] = y['Name'].copy()
new_df2['Media'] = y['Media'].copy()
new_df2['Resolution'] = y['Resolution'].copy()

# print(new_df2)
new_df2.to_excel('extract_condensed.xlsx') # create condensed version of extract

# COMPARE WORKBOOKS (raw_file_condensed.xlsx and extract_condensed.xlsx)
import openpyxl
from openpyxl.styles import PatternFill

workbook1 = openpyxl.load_workbook('raw_file_condensed.xlsx')
workbook2 = openpyxl.load_workbook('extract_condensed.xlsx')

cell_background_fill = PatternFill(start_color="E2BBBB", end_color="E2BBBB", fill_type="solid")
# rename the sheet to 'sheet4' when generated
workbook1_sheet = workbook1['Sheet1']
workbook1_sheet.title = 'Sheet4'
workbook1_sheet.delete_cols(1) # remove array column automatically added in column a
workbook1.save('raw_file_condensed.xlsx')
# rename sheet to 'sheet3' when generated
workbook2_sheet = workbook2['Sheet1']
workbook2_sheet.title = 'Sheet3'
workbook2_sheet.delete_cols(1)  # remove array column automatically added in column a
workbook2.save('extract_condensed.xlsx')
# for loop to iterate around condensed raw file
for row in workbook1_sheet.iter_rows():
    for cell in row:
        current_cell_value = cell.value
        cell_location = cell.coordinate
        # if the value in the cell of workbook2 does not match workbook2
        try:
            if current_cell_value != workbook2_sheet[cell_location].value:
                cell.fill = cell_background_fill
        except Exception as e:
                print(e)

# change name of sheet in results workbook from 'sheet4' to 'results'
workbook1_sheet = workbook1['Sheet4']
workbook1_sheet.title = 'Results'

# auto size column widths
workbook1_sheet.column_dimensions["A"].width = 40.00
workbook1_sheet.column_dimensions["B"].width = 9.14
workbook1_sheet.column_dimensions["C"].width = 9.86

workbook1.save('results.xlsx') # create comparison of sheets in "results" workbook

'''
# prevent results workbook from being edited
workbook1_sheet.protection.sheet = True
workbook1.save("results.xlsx")
'''
